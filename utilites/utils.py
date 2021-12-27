import inspect
import logging
import softest
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

class Utils(softest.TestCase):
    def assert_value_in_list(self, listVal, textVal):
        for val in listVal:
            self.soft_assert(self.assertEqual, val.text, textVal)
            if val.text == textVal:
                print("Test pass")
            else:
                print("Test Fail")
        self.assert_all()

    @staticmethod
    def loggen(logLevel=logging.DEBUG):
        # Set class/method name from where ít called
        logger_name = inspect.stack()[1][3]

        # Create Logger
        logger = logging.getLogger(logger_name)
        logger.setLevel(logLevel)

        # Create console handler or file handler and set the log level
        fh = logging.FileHandler("Automation.Log", mode='a')

        # Create Formatter
        formatter = logging.Formatter('%(asctime)s - %(levelname)s -%(name)s : %(message)s', datefmt='%d/%m/%Y %I:%M:%S %p')

        # Add formatter to console or file handler
        fh.setFormatter(formatter)

        # Add console handler to logger
        logger.addHandler(fh)
        return logger

    @staticmethod
    def read_data_from_exel(file_name, sheet):
        dataList = []
        wb = load_workbook(filename=file_name)
        sh: Worksheet = wb[sheet]
        row_ct = sh.max_row
        col_ct = sh.max_column

        for i in range(2, row_ct+1):
            row = []
            for j in range(2, col_ct + 1):
                row.append(sh.cell(row=i, column=j).value)
            dataList.append(row)
        wb.close()
        return dataList

# data = Utils()
# data_get = data.read_data_from_exel("E:\\Auto_WorkingTesting\\ATFDemo\\testdata\\testdata.xlsx", "Sheet1")
# print(data_get)